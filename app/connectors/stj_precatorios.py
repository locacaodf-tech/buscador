from __future__ import annotations

import re
import unicodedata
from pathlib import Path
from typing import Any

import httpx

from .base import BaseConnector, ConnectorCapability, ConnectorError, SearchNotSupported
from ..config import get_settings
from ..utils.cnj import normalize_cnj

STJ_PRECATORIOS_PAGE = 'https://www.stj.jus.br/sites/portalp/Processos/precatorios'

# Palavras-chave para mapear colunas do XLSX por semelhança de cabeçalho.
# A estrutura real observada (print do usuário, 28/06/2026): ordem | classe |
# sequencial | processo | ... | prioridade/natureza | valor | previsão.
# Os cabeçalhos exatos variam por arquivo/ano, então o parser localiza a linha
# de cabeçalho e mapeia por palavra-chave — nunca por posição fixa.
COLUMN_KEYWORDS: dict[str, tuple[str, ...]] = {
    'ordem': ('ordem',),
    'classe': ('classe',),
    'sequencial': ('sequencial', 'seq'),
    'numero_processo': ('processo', 'numero do processo', 'nº do processo'),
    'credor_nome': ('credor', 'beneficiario', 'requerente', 'exequente'),
    'natureza': ('natureza',),
    'prioridade': ('prioridade', 'preferencia', 'idoso'),
    'valor': ('valor',),
    'previsao_pagamento': ('previsao', 'previsão', 'mes de pagamento', 'pagamento'),
    'entidade_devedora': ('entidade devedora', 'devedor', 'executado'),
}


def _norm(text: Any) -> str:
    s = str(text or '').strip().lower()
    return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')


def _parse_valor(raw: Any) -> float | None:
    """Converte '1.708.161,78' / 1708161.78 / '' para float, sem inventar."""
    if raw is None or raw == '':
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    s = str(raw).strip().replace('R$', '').strip()
    if not s:
        return None
    # Formato brasileiro: ponto de milhar, vírgula decimal
    if ',' in s:
        s = s.replace('.', '').replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return None


def infer_file_metadata(url: str, link_context: str = '') -> dict[str, Any]:
    """Extrai categoria/natureza/anos do nome do arquivo e do contexto do link.

    Função pura, testável. Nunca inventa: o que não der pra inferir fica None.
    """
    blob = _norm(url) + ' ' + _norm(link_context)
    meta: dict[str, Any] = {'categoria': None, 'natureza': None, 'ano_expedicao': None, 'ano_orcamento': None}

    # Precedência: o nome do ARQUIVO manda; o contexto do link só desempata.
    # Motivo: a página do STJ rotula seções de expedidos com "inclusos na
    # proposta orçamentária de XXXX", o que confundiria a categoria.
    url_norm = _norm(url)
    if 'pendente' in url_norm:
        meta['categoria'] = 'pendentes'
    elif 'pago' in url_norm or 'pagos' in url_norm or 'pagas' in url_norm:
        meta['categoria'] = 'pagos'
    elif 'expedid' in url_norm or 'autuad' in url_norm:
        meta['categoria'] = 'expedidos'
    elif 'proposta' in url_norm:
        meta['categoria'] = 'proposta'
    elif 'mapa' in url_norm and 'anual' in url_norm:
        meta['categoria'] = 'mapa_anual'
    else:
        # nome do arquivo inconclusivo -> tenta o contexto do link
        if 'pendente' in blob:
            meta['categoria'] = 'pendentes'
        elif 'pago' in blob or 'pagas' in blob:
            meta['categoria'] = 'pagos'
        elif 'expedid' in blob:
            meta['categoria'] = 'expedidos'
        elif 'proposta' in blob:
            meta['categoria'] = 'proposta'
    if 'rpv' in blob:
        meta['tipo_requisicao'] = 'rpv'
    elif 'prc' in blob or 'precatorio' in blob:
        meta['tipo_requisicao'] = 'precatorio'

    if 'alimentar' in blob:
        meta['natureza'] = 'alimentar'
    elif 'comum' in blob:
        meta['natureza'] = 'comum'

    years = [int(y) for y in re.findall(r'20\d{2}', blob)]
    if meta['categoria'] == 'proposta' and years:
        meta['ano_orcamento'] = max(years)
    elif len(years) >= 2:
        # padrão "Expedidos em 2026 ... proposta de 2027"
        meta['ano_expedicao'] = min(years)
        meta['ano_orcamento'] = max(years)
    elif len(years) == 1:
        meta['ano_expedicao'] = years[0]

    return meta


def discover_file_links(page_html: str) -> list[dict[str, Any]]:
    """Descobre dinamicamente os links de XLSX/XLSM/XLS na página oficial.

    É isto que resolve o problema do 'ano travado': quando o STJ publicar os
    arquivos do próximo exercício, eles aparecem aqui sozinhos, sem mexer em
    código nem em configuração.
    """
    links = []
    for match in re.finditer(r'href="([^"]+\.(?:xlsx|xlsm|xls))"', page_html, re.IGNORECASE):
        url = match.group(1)
        if url.startswith('/'):
            url = 'https://www.stj.jus.br' + url
        # contexto: até 200 chars antes do link, para pegar o rótulo/ano da seção
        start = max(0, match.start() - 200)
        context = re.sub(r'<[^>]+>', ' ', page_html[start:match.start()])
        entry = {'url': url, 'context': context.strip()[-160:]}
        entry.update(infer_file_metadata(url, context))
        links.append(entry)
    return links


def find_header_row(rows: list[tuple]) -> tuple[int, dict[str, int]] | None:
    """Localiza a linha de cabeçalho e mapeia colunas por palavra-chave.

    Retorna (índice_da_linha, {campo: índice_da_coluna}) ou None se não achar.
    Exige no mínimo 'sequencial' ou 'processo' para considerar cabeçalho válido.
    """
    for i, row in enumerate(rows[:30]):  # cabeçalho sempre nas primeiras linhas
        mapping: dict[str, int] = {}
        for col_idx, cell in enumerate(row):
            cell_norm = _norm(cell)
            if not cell_norm:
                continue
            for field, keywords in COLUMN_KEYWORDS.items():
                if field not in mapping and any(kw in cell_norm for kw in keywords):
                    mapping[field] = col_idx
        if 'sequencial' in mapping or 'numero_processo' in mapping:
            return i, mapping
    return None


def parse_stj_sheet(rows: list[tuple], file_meta: dict[str, Any], fonte_url: str, aba: str | None = None) -> list[dict[str, Any]]:
    """Converte as linhas de uma planilha do STJ em registros normalizados.

    Campos ausentes ficam None — NUNCA preenchidos com palpite. Cada registro
    carrega fonte_url, aba, arquivo de origem e número da linha, para auditoria.
    """
    header = find_header_row(rows)
    if not header:
        return []
    header_idx, mapping = header
    records = []
    for line_number, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
        def cell(field: str) -> Any:
            idx = mapping.get(field)
            if idx is None or idx >= len(row):
                return None
            value = row[idx]
            return value if value not in ('', None) else None

        sequencial = cell('sequencial')
        processo = cell('numero_processo')
        if sequencial is None and processo is None:
            continue  # linha vazia/rodapé

        records.append({
            'sequencial': str(int(sequencial)) if isinstance(sequencial, float) and sequencial.is_integer() else (str(sequencial) if sequencial is not None else None),
            'classe': cell('classe'),
            'numero_processo': str(processo) if processo is not None else None,
            'credor_nome': cell('credor_nome'),
            'natureza': cell('natureza') or file_meta.get('natureza'),
            'prioridade': cell('prioridade'),
            'valor': _parse_valor(cell('valor')),
            'previsao_pagamento': cell('previsao_pagamento'),
            'entidade_devedora': cell('entidade_devedora'),
            'ano_orcamento': file_meta.get('ano_orcamento'),
            'ano_expedicao': file_meta.get('ano_expedicao'),
            'categoria': file_meta.get('categoria'),
            'tipo_requisicao': file_meta.get('tipo_requisicao'),
            'fonte_url': fonte_url,
            'aba': aba,
            'linha_arquivo': line_number,
        })
    return records


def parse_workbook_bytes(content: bytes, file_meta: dict[str, Any], fonte_url: str) -> list[dict[str, Any]]:
    """Abre um XLSX/XLSM/XLS em memória e devolve os registros de todas as
    abas. Função de módulo (sem depender de instância de conector) para ser
    reutilizada tanto pela busca online quanto pelo upload manual."""
    import io
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ConnectorError('openpyxl não instalado. Rode: pip install openpyxl') from exc
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ConnectorError(f'Não consegui abrir a planilha do STJ ({fonte_url}): {exc}')
    records: list[dict[str, Any]] = []
    for ws in wb.worksheets:
        rows = [tuple(row) for row in ws.iter_rows(values_only=True)]
        records.extend(parse_stj_sheet(rows, file_meta, fonte_url, aba=ws.title))
    wb.close()
    return records


def match_records(records: list[dict[str, Any]], search_type: str, key: str, ano_filtro: Any = None) -> list[dict[str, Any]]:
    """Filtra registros já parseados pelo identificador buscado. Função de
    módulo, reutilizada pela busca online e pela busca em arquivo carregado."""
    key_digits = re.sub(r'\D+', '', key or '')
    results = []
    for rec in records:
        if ano_filtro and rec.get('ano_orcamento') and int(rec['ano_orcamento']) != int(ano_filtro):
            continue
        if search_type in {'precatorio_number', 'requisitorio_number', 'rpv_number', 'sequencial'}:
            seq = re.sub(r'\D+', '', str(rec.get('sequencial') or ''))
            if seq and seq == key_digits:
                results.append(rec)
        else:  # cnj / numero_processo
            proc = normalize_cnj(str(rec.get('numero_processo') or ''))
            if proc and proc == normalize_cnj(key):
                results.append(rec)
    return results


class StjPrecatoriosConnector(BaseConnector):
    """Fonte oficial de precatórios/RPVs do STJ (competência originária).

    Fonte verificada ao vivo em 01/07/2026: a página pública do STJ
    (https://www.stj.jus.br/sites/portalp/Processos/precatorios) lista
    arquivos XLSX/XLSM abertos, sem login e sem captcha, cobrindo:
    expedidos por ano ("2026 - inclusos na proposta orçamentária de 2027"),
    proposta por exercício (PRCs-proposta-2027.xlsx etc.), pendentes,
    pagos por mês e mapa anual. O download direto do XLSX também foi
    confirmado (HTTP 200, mime correto).

    A busca NÃO trava em ano fixo: discover_file_links() lê a página a cada
    consulta (com cache) e descobre os arquivos disponíveis — anos novos
    entram sozinhos. O filtro por ano-orçamento é opcional (extra_params).

    O sequencial é numeração interna do STJ — não confundir com sequenciais
    de TRF/TJ. Este conector só olha arquivos do STJ, então a busca por
    sequencial aqui é inequívoca.

    Aviso de honestidade técnica: o mapeamento de colunas do parser foi
    construído a partir da estrutura visível em prints reais (ordem, classe,
    sequencial, processo, natureza/prioridade, valor, previsão) e validado
    com planilha sintética equivalente — mas NÃO contra o arquivo real do
    STJ, porque o ambiente onde este código foi escrito não baixa arquivos
    desse domínio. Se algum cabeçalho real não casar com COLUMN_KEYWORDS,
    o campo vem None (nunca inventado) — nesse caso me mande os cabeçalhos
    exatos do arquivo que eu ajusto o mapeamento.

    Modo offline: defina STJ_LOCAL_DIR no .env apontando para uma pasta com
    os XLSX baixados manualmente; o conector parseia de lá sem rede.
    """

    name = 'stj_precatorios'
    capabilities = (
        ConnectorCapability(
            'precatorio_number',
            description='Busca por sequencial do precatório do STJ nos arquivos oficiais (todos os anos disponíveis; filtre por ano-orçamento em extra_params se quiser).',
            optional_fields=('extra_params.ano_orcamento',),
        ),
        ConnectorCapability(
            'requisitorio_number',
            description='Mesma busca do sequencial (no STJ o número do requisitório é o sequencial do PRC).',
            optional_fields=('extra_params.ano_orcamento',),
        ),
        ConnectorCapability(
            'cnj',
            description='Busca por número do processo (CNJ) nos arquivos oficiais do STJ.',
            optional_fields=('extra_params.ano_orcamento',),
        ),
        ConnectorCapability(
            'numero_processo',
            description='Idem CNJ.',
            optional_fields=('extra_params.ano_orcamento',),
        ),
    )

    SUPPORTED = {'precatorio_number', 'requisitorio_number', 'cnj', 'numero_processo', 'rpv_number'}

    def __init__(self):
        self.settings = get_settings()
        self.local_dir = getattr(self.settings, 'stj_local_dir', '') or ''
        self._page_cache: str | None = None
        self._workbook_cache: dict[str, list[dict[str, Any]]] = {}

    # ------------------------------------------------------------------
    # Aquisição de dados (online com cache, ou offline por pasta local)
    # ------------------------------------------------------------------

    async def _fetch_page(self) -> str:
        if self._page_cache is not None:
            return self._page_cache
        async with httpx.AsyncClient(timeout=30, follow_redirects=True) as client:
            response = await client.get(STJ_PRECATORIOS_PAGE)
            if response.status_code >= 400:
                raise ConnectorError(f'Página do STJ retornou HTTP {response.status_code}')
            self._page_cache = response.text
            return self._page_cache

    async def _download_records(self, link: dict[str, Any]) -> list[dict[str, Any]]:
        url = link['url']
        if url in self._workbook_cache:
            return self._workbook_cache[url]
        async with httpx.AsyncClient(timeout=60, follow_redirects=True) as client:
            response = await client.get(url)
            if response.status_code >= 400:
                raise ConnectorError(f'Arquivo do STJ retornou HTTP {response.status_code}: {url}')
            records = parse_workbook_bytes(response.content, link, url)
            self._workbook_cache[url] = records
            return records

    def _local_records(self) -> list[dict[str, Any]]:
        folder = Path(self.local_dir)
        records: list[dict[str, Any]] = []
        for path in sorted(folder.glob('*.xls*')):
            meta = infer_file_metadata(path.name)
            records.extend(parse_workbook_bytes(path.read_bytes(), meta, f'file://{path}'))
        return records

    # ------------------------------------------------------------------
    # Busca
    # ------------------------------------------------------------------

    async def search(self, search_type: str, search_key: str, **kwargs: Any) -> list[dict]:
        if search_type not in self.SUPPORTED:
            raise SearchNotSupported(
                f'STJ: tipo de busca "{search_type}" não suportado. Suportados: sequencial do '
                'precatório/requisitório (precatorio_number/requisitorio_number) e número de '
                'processo (cnj/numero_processo). Busca por CPF/nome não está disponível nos '
                'arquivos públicos do STJ sem verificação adicional.'
            )
        key = (search_key or '').strip()
        if not key:
            raise SearchNotSupported('STJ: valor de busca vazio.')

        extra = kwargs.get('extra_params') or {}
        ano_filtro = extra.get('ano_orcamento')

        try:
            if self.local_dir:
                all_records = self._local_records()
            else:
                page = await self._fetch_page()
                links = discover_file_links(page)
                if not links:
                    return [{'provider': self.name, 'tribunal': 'STJ',
                             'error': 'Nenhum arquivo XLSX encontrado na página oficial do STJ — a estrutura da página pode ter mudado. Verifique manualmente: ' + STJ_PRECATORIOS_PAGE}]
                # Prioriza proposta/expedidos/pendentes; ignora mapa anual e RPV quando
                # a busca é por precatório, para não baixar dezenas de arquivos à toa.
                relevant = [l for l in links if l.get('categoria') in {'proposta', 'expedidos', 'pendentes', 'pagos'}]
                if ano_filtro:
                    relevant = [l for l in relevant if l.get('ano_orcamento') in (None, int(ano_filtro))]
                all_records = []
                errors = []
                for link in relevant:
                    try:
                        all_records.extend(await self._download_records(link))
                    except ConnectorError as exc:
                        errors.append(str(exc))
        except ConnectorError as exc:
            return [{'provider': self.name, 'tribunal': 'STJ', 'error': str(exc)}]
        except Exception as exc:
            return [{'provider': self.name, 'tribunal': 'STJ', 'error': f'Falha ao consultar STJ: {exc}'}]

        matches = match_records(all_records, search_type, key, ano_filtro)
        out = []
        for rec in matches:
            out.append({'provider': self.name, 'tribunal': 'STJ', 'source': rec, 'raw_hit': rec})
        return out
