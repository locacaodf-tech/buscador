from __future__ import annotations

import io
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from ..config import get_settings
from ..connectors.stj_precatorios import (
    find_header_row,
    infer_file_metadata,
    match_records,
    parse_stj_sheet,
    parse_workbook_bytes,
)

UPLOAD_DIR = Path('data/stj_uploads')
ALLOWED_EXTENSIONS = {'.xlsx', '.xlsm', '.xls'}
MAX_UPLOAD_BYTES = 20 * 1024 * 1024  # 20MB — generoso para os arquivos do STJ (poucos MB tipicamente)

TIMESTAMP_FMT = '%Y%m%dT%H%M%SZ'
SEP = '__'


class InvalidUpload(ValueError):
    pass


def _sanitize_filename(name: str) -> str:
    """Nunca confia no nome enviado pelo cliente: tira qualquer caminho
    (impede path traversal tipo '../../etc/passwd') e caracteres fora de um
    conjunto seguro."""
    name = Path(name).name
    name = re.sub(r'[^A-Za-z0-9._-]+', '_', name)
    return name or 'arquivo.xlsx'


def upload_dir() -> Path:
    # Permite produção persistente via STJ_UPLOAD_DIR=/data/stj_uploads.
    # Se a variável não estiver definida, preserva o comportamento local/testes.
    configured = getattr(get_settings(), 'stj_upload_dir', '')
    return Path(configured) if configured else UPLOAD_DIR


def ensure_upload_dir() -> Path:
    directory = upload_dir()
    directory.mkdir(parents=True, exist_ok=True)
    return directory


@dataclass
class UploadedFileInfo:
    original_filename: str
    stored_filename: str
    stored_path: Path
    uploaded_at: str  # ISO 8601


def _uploaded_at_from_stored_name(stored_filename: str, fallback_path: Path | None = None) -> str:
    if SEP in stored_filename:
        prefix = stored_filename.split(SEP, 1)[0]
        try:
            dt = datetime.strptime(prefix, TIMESTAMP_FMT).replace(tzinfo=timezone.utc)
            return dt.isoformat()
        except ValueError:
            pass
    if fallback_path is not None and fallback_path.exists():
        return datetime.fromtimestamp(fallback_path.stat().st_mtime, tz=timezone.utc).isoformat()
    return datetime.now(timezone.utc).isoformat()


def save_uploaded_file(original_filename: str, content: bytes) -> UploadedFileInfo:
    ext = Path(original_filename or '').suffix.lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise InvalidUpload(f'Extensão "{ext}" não permitida. Envie .xlsx, .xlsm ou .xls.')
    if not content:
        raise InvalidUpload('Arquivo vazio.')
    if len(content) > MAX_UPLOAD_BYTES:
        raise InvalidUpload(f'Arquivo maior que o limite de {MAX_UPLOAD_BYTES // (1024 * 1024)}MB.')

    ensure_upload_dir()
    safe_name = _sanitize_filename(original_filename)
    timestamp = datetime.now(timezone.utc).strftime(TIMESTAMP_FMT)
    stored_filename = f'{timestamp}{SEP}{safe_name}'
    stored_path = ensure_upload_dir() / stored_filename
    stored_path.write_bytes(content)
    return UploadedFileInfo(
        original_filename=original_filename,
        stored_filename=stored_filename,
        stored_path=stored_path,
        uploaded_at=_uploaded_at_from_stored_name(stored_filename, stored_path),
    )


def list_uploaded_files() -> list[UploadedFileInfo]:
    ensure_upload_dir()
    out = []
    directory = ensure_upload_dir()
    for path in sorted(directory.glob('*.xls*')):
        stored_filename = path.name
        original = stored_filename.split(SEP, 1)[1] if SEP in stored_filename else stored_filename
        out.append(UploadedFileInfo(
            original_filename=original,
            stored_filename=stored_filename,
            stored_path=path,
            uploaded_at=_uploaded_at_from_stored_name(stored_filename, path),
        ))
    return out


def inspect_workbook(content: bytes, original_filename: str = '') -> dict[str, Any]:
    """Raio-x do arquivo recém-enviado: abas, cabeçalho encontrado, campos
    detectados e quantas linhas foram lidas — sem precisar de search_key
    ainda. Isso é o que o endpoint de upload devolve."""
    wb_bytes = io.BytesIO(content)
    from openpyxl import load_workbook
    wb = load_workbook(wb_bytes, read_only=True, data_only=True)
    file_meta = infer_file_metadata(original_filename)
    sheets_info = []
    total_registros = 0
    for ws in wb.worksheets:
        rows = [tuple(row) for row in ws.iter_rows(values_only=True)]
        header = find_header_row(rows)
        campos = sorted(header[1].keys()) if header else []
        records = parse_stj_sheet(rows, file_meta, 'upload://pending', aba=ws.title) if header else []
        total_registros += len(records)
        sheets_info.append({
            'aba': ws.title,
            'linhas_totais': len(rows),
            'linha_cabecalho': (header[0] + 1) if header else None,
            'campos_detectados': campos,
            'registros_lidos': len(records),
        })
    wb.close()
    return {
        'categoria_detectada': file_meta.get('categoria'),
        'natureza_detectada': file_meta.get('natureza'),
        'ano_orcamento_detectado': file_meta.get('ano_orcamento'),
        'abas': sheets_info,
        'total_registros': total_registros,
    }


def search_uploaded_files(search_type: str, search_key: str, ano_orcamento: int | None = None) -> dict[str, Any]:
    """Busca nos arquivos que já foram enviados via upload. Se nenhum
    arquivo foi carregado ainda, devolve status 'not_loaded' — nunca cai
    para dado sintético ou inventado."""
    files = list_uploaded_files()
    if not files:
        return {
            'status': 'not_loaded',
            'message': 'Arquivo oficial do STJ ainda não carregado.',
            'files_loaded': [],
            'total': 0,
            'results': [],
        }

    consultado_em = datetime.now(timezone.utc).isoformat()
    results: list[dict[str, Any]] = []
    files_summary: list[dict[str, Any]] = []

    for info in files:
        content = info.stored_path.read_bytes()
        file_meta = infer_file_metadata(info.original_filename)
        try:
            records = parse_workbook_bytes(content, file_meta, f'upload:{info.original_filename}')
        except Exception as exc:
            files_summary.append({
                'arquivo_original': info.original_filename,
                'uploaded_at': info.uploaded_at,
                'erro': str(exc),
            })
            continue

        # parse_workbook_bytes não sabe o nome do arquivo original para o
        # rótulo de 'aba' por planilha — os registros já carregam 'aba' do
        # parse_stj_sheet, só precisamos anotar a origem do upload.
        matches = match_records(records, search_type, search_key, ano_orcamento)
        for rec in matches:
            rec_out = dict(rec)
            rec_out.update({
                'fonte': 'STJ - XLSX oficial carregado manualmente',
                'arquivo_original': info.original_filename,
                'arquivo_salvo': info.stored_filename,
                'uploaded_at': info.uploaded_at,
                'consultado_em': consultado_em,
            })
            results.append(rec_out)

        files_summary.append({
            'arquivo_original': info.original_filename,
            'uploaded_at': info.uploaded_at,
            'registros': len(records),
        })

    return {
        'status': 'completed',
        'message': f'{len(results)} registro(s) encontrado(s) em {len(files)} arquivo(s) carregado(s).',
        'files_loaded': files_summary,
        'total': len(results),
        'results': results,
    }
